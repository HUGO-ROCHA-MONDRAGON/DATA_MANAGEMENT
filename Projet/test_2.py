import pandas as pd
from openpyxl import load_workbook, Workbook
from openpyxl.styles import Font, PatternFill, Border, Side, Alignment
from openpyxl.utils import get_column_letter
from datetime import date, timedelta

# File paths
excel_ratings = "C:/Users/h24826/BNP Paribas/GFI - GCC - Structured Credit - Documents/Portfolio Management/Hugo/Ratings/RATINGS.xlsx"
excel_output = "C:/Users/h24826/BNP Paribas/GFI - GCC - Structured Credit - Documents/Portfolio Management/Hugo/Ratings/RATINGS_output.xlsx"
output_wsname = "ratings"

# Sheet pairs
sheet_pairs = [
    ("T_MINUS_ONE_A", "T_MINUS_TWO_A"),
    ("T_MINUS_ONE_B", "T_MINUS_TWO_B"),
    ("T_MINUS_ONE_C", "T_MINUS_TWO_C")
]

# Labels for custom headers
label_map = {
    "T_MINUS_ONE_A": "FLABSA",
    "T_MINUS_ONE_B": "FLABSB",
    "T_MINUS_ONE_C": "FLABSC"
}

# Dates
today = date.today()
minus_1 = today - timedelta(days=1)
minus_2 = today - timedelta(days=2)

# Style setup
bold_font = Font(bold=True)
italic_font = Font(italic=True)
header_fill = PatternFill(start_color="D9E1F2", end_color="D9E1F2", fill_type="solid")
upgrade_fill = PatternFill(start_color="C6EFCE", end_color="C6EFCE", fill_type="solid")
downgrade_fill = PatternFill(start_color="FFC7CE", end_color="FFC7CE", fill_type="solid")
white_fill = PatternFill(start_color="FFFFFF", end_color="FFFFFF", fill_type="solid")
thin_border = Border(
    left=Side(style="thin"), right=Side(style="thin"),
    top=Side(style="thin"), bottom=Side(style="thin")
)
center_align = Alignment(horizontal="center", vertical="center")

# Rating scales (lower = better)
MOODYS_SCALE = {
    "Aaa": 1, "Aa1": 2, "Aa2": 3, "Aa3": 4, "A1": 5, "A2": 6, "A3": 7,
    "Baa1": 8, "Baa2": 9, "Baa3": 10, "Ba1": 11, "Ba2": 12, "Ba3": 13,
    "B1": 14, "B2": 15, "B3": 16, "Caa1": 17, "Caa2": 18, "Caa3": 19,
    "Ca": 20, "C": 21
}
FITCH_SCALE = S_AND_P_SCALE = {
    "AAA": 1, "AA+": 2, "AA": 3, "AA-": 4, "A+": 5, "A": 6, "A-": 7,
    "BBB+": 8, "BBB": 9, "BBB-": 10, "BB+": 11, "BB": 12, "BB-": 13,
    "B+": 14, "B": 15, "B-": 16, "CCC+": 17, "CCC": 18, "CCC-": 19,
    "CC": 20, "C": 21, "D": 22
}

# Rating change detection
def get_rating_change(row, agency):
    old_rating = row[f"{agency}_minus2"]
    new_rating = row[f"{agency}_minus1"]
    if pd.isna(old_rating) or pd.isna(new_rating):
        return ""
    if old_rating == new_rating:
        return ""
    return f"{old_rating} → {new_rating}"

# Open/create workbook
try:
    wb = load_workbook(excel_output)
    if output_wsname in wb.sheetnames:
        del wb[output_wsname]
    ws = wb.create_sheet(output_wsname)
except FileNotFoundError:
    wb = Workbook()
    ws = wb.active
    ws.title = output_wsname

start_row = 2

# Loop through each sheet pair
for sheet1, sheet2 in sheet_pairs:
    t1 = pd.read_excel(excel_ratings, sheet_name=sheet1)
    t2 = pd.read_excel(excel_ratings, sheet_name=sheet2)

    for df in [t1, t2]:
        df["ISIN"] = df["ISIN"].astype(str).str.strip()
        df["Security_description"] = df["Security_description"].astype(str).str.strip()

    # Rename rating and price columns in T-1
    t1 = t1.rename(columns={
        "Moodys": "Moodys_minus1",
        "Fitch": "Fitch_minus1",
        "S&P": "S&P_minus1",
        "Close price": "Close_price"
    })

    # Rename only ratings in T-2
    t2 = t2.rename(columns={
        "Moodys": "Moodys_minus2",
        "Fitch": "Fitch_minus2",
        "S&P": "S&P_minus2"
    })

    # Merge
    df = t1.merge(t2, on=["ISIN", "Security_description"], how="inner")

    # Compute changes
    for agency in ["Moodys", "Fitch", "S&P"]:
        df[f"{agency}_change"] = df.apply(lambda row: get_rating_change(row, agency), axis=1)

    # Keep only rows with at least one change
    changes = df[
        (df[["Moodys_change", "Fitch_change", "S&P_change"]] != "").any(axis=1)
    ][["ISIN", "Security_description", "Close_price", "Moodys_change", "Fitch_change", "S&P_change"]]

    # Drop duplicate ISINs
    changes = changes.drop_duplicates(subset=["ISIN"])

    # Write section title
    label = label_map.get(sheet1, "Unknown")
    header_text = f"{label}: {minus_1.strftime('%d/%m/%Y')} vs {minus_2.strftime('%d/%m/%Y')}"
    ws.cell(row=start_row, column=1, value=header_text).font = bold_font
    start_row += 1

    if changes.empty:
        cell = ws.cell(row=start_row, column=1, value="No changes")
        cell.font = italic_font
        cell.fill = white_fill
        start_row += 2
    else:
        headers = ["ISIN", "Security_description", "Close_price", "Moodys_change", "Fitch_change", "S&P_change"]
        for col_idx, header in enumerate(headers, start=1):
            cell = ws.cell(row=start_row, column=col_idx, value=header)
            cell.font = bold_font
            cell.fill = header_fill
            cell.border = thin_border
            cell.alignment = center_align
        start_row += 1

        for row in changes.itertuples(index=False):
            for col_idx, value in enumerate(row, start=1):
                cell = ws.cell(row=start_row, column=col_idx, value=value)
                cell.border = thin_border
                cell.alignment = center_align

                # Conditional coloring for rating change cells
                if col_idx >= 4 and isinstance(value, str) and "→" in value:
                    left, right = value.split("→")
                    left = left.strip().replace("(", "").replace(")", "")
                    right = right.strip().replace("(", "").replace(")", "")
                    scale = None
                    if col_idx == 4:
                        scale = MOODYS_SCALE
                    elif col_idx == 5:
                        scale = FITCH_SCALE
                    elif col_idx == 6:
                        scale = S_AND_P_SCALE
                    score_left = scale.get(left)
                    score_right = scale.get(right)
                    if score_left is not None and score_right is not None:
                        if score_left < score_right:
                            cell.fill = downgrade_fill
                        elif score_left > score_right:
                            cell.fill = upgrade_fill
                        else:
                            cell.fill = white_fill
                    else:
                        cell.fill = white_fill
                else:
                    cell.fill = white_fill
            start_row += 1
        start_row += 1

# Auto column widths
for column_cells in ws.columns:
    max_len = 0
    col_letter = get_column_letter(column_cells[0].column)
    for cell in column_cells:
        if cell.value:
            max_len = max(max_len, len(str(cell.value)))
    ws.column_dimensions[col_letter].width = max_len + 2

# Save
wb.save(excel_output)
print(f"Data inserted into '{output_wsname}' in {excel_output}")